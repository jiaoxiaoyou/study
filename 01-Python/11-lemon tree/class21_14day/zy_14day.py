# @Author : jiaojie
# @CreateDate : 2020/5/30 21:22
# @Description : 
# @UpdateAuthor : 
# @LastUpdateTime : 
# @UpdateDescription :

"""
workbook：工作簿对象
sheet：表单对象
cell：表格对象
"""

import openpyxl

workbook = openpyxl.load_workbook('case.xlsx')
sheet = workbook['Sheet1']
max_row = sheet.max_row
max_column = sheet.max_column
print("max_row:{},max_column:{}".format(max_row,max_column))
cases = []
for r in range(1, max_row+1):
    if r == 1:
        titles = []
        for c in range(1, max_column+1):
            data = sheet.cell(row=r, column=c).value
            titles.append(data)
    else:
        datas = []
        for c in range(1, max_column+1):
            data = sheet.cell(row=r, column=c).value
            datas.append(data)
            #print(titles, datas)
        case_zip = zip(titles, datas)
        case = dict(case_zip)
        cases.append(case)
print(cases)

